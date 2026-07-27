import io
from collections import Counter
from typing import List

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import openpyxl

from app.database import get_db
from app.deps import require_roles
from app.models.user import User, UserRole
from app.models.skin_profile import SkinProfile
from app.models.progress import ProgressLog
from app.models.recommendation import Recommendation
from app.schemas.user import UserOut

router = APIRouter(prefix="/api/admin", tags=["Admin"], dependencies=[Depends(require_roles(UserRole.admin))])


@router.get("/users", response_model=List[UserOut])
def list_all_users(db: Session = Depends(get_db)):
    return db.query(User).all()


@router.post("/users/{user_id}/deactivate")
def deactivate_user(user_id: str, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if user:
        user.is_active = False
        db.commit()
    return {"ok": True}


@router.post("/users/{user_id}/activate")
def activate_user(user_id: str, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if user:
        user.is_active = True
        db.commit()
    return {"ok": True}


@router.get("/analytics")
def platform_analytics(db: Session = Depends(get_db)):
    role_counts = {role.value: db.query(User).filter(User.role == role).count() for role in UserRole}

    concern_counter = Counter()
    for profile in db.query(SkinProfile).all():
        for concern in (profile.skin_concerns or []):
            concern_counter[concern] += 1
    top_concerns = [{"concern": c, "count": n} for c, n in concern_counter.most_common(10)]

    scores = []
    user_ids_with_logs = [row[0] for row in db.query(ProgressLog.user_id).distinct().all()]
    for uid in user_ids_with_logs:
        latest = (
            db.query(ProgressLog)
            .filter(ProgressLog.user_id == uid)
            .order_by(ProgressLog.log_date.desc())
            .first()
        )
        if latest and latest.skin_health_score is not None:
            scores.append(latest.skin_health_score)
    average_skin_health_score = round(sum(scores) / len(scores), 2) if scores else None

    return {
        "role_counts": role_counts,
        "top_concerns": top_concerns,
        "average_skin_health_score": average_skin_health_score,
        "users_with_progress_logs": len(scores),
    }


@router.get("/recommendations")
def all_recommendations(db: Session = Depends(get_db)):
    recs = db.query(Recommendation).order_by(Recommendation.created_at.desc()).all()
    result = []
    for r in recs:
        client = db.query(User).filter(User.id == r.client_id).first()
        author = db.query(User).filter(User.id == r.author_id).first()
        result.append({
            "id": str(r.id),
            "client_name": client.full_name if client else "Unknown",
            "author_name": author.full_name if author else "Unknown",
            "author_role": r.author_role,
            "note": r.note,
            "created_at": r.created_at,
        })
    return result


@router.get("/reports/excel")
def platform_excel_report(db: Session = Depends(get_db)):
    wb = openpyxl.Workbook()

    ws_users = wb.active
    ws_users.title = "Users"
    ws_users.append(["Name", "Email", "Role", "Active"])
    for u in db.query(User).all():
        ws_users.append([u.full_name, u.email, u.role.value, u.is_active])

    ws_profiles = wb.create_sheet("Skin Profiles")
    ws_profiles.append(["User Email", "Skin Type", "Age Group", "Concerns"])
    for p in db.query(SkinProfile).all():
        owner = db.query(User).filter(User.id == p.user_id).first()
        ws_profiles.append([
            owner.email if owner else "-",
            p.skin_type,
            p.age_group,
            ", ".join(p.skin_concerns or []),
        ])

    ws_recs = wb.create_sheet("Recommendations")
    ws_recs.append(["Client Email", "Author", "Role", "Note", "Date"])
    for r in db.query(Recommendation).all():
        client = db.query(User).filter(User.id == r.client_id).first()
        author = db.query(User).filter(User.id == r.author_id).first()
        ws_recs.append([
            client.email if client else "-",
            author.full_name if author else "-",
            r.author_role,
            r.note,
            r.created_at.strftime("%Y-%m-%d"),
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=platform_report.xlsx"},
    )