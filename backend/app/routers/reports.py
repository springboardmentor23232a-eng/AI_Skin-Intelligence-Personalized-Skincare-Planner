import io
import os
from datetime import datetime, timedelta
from typing import Any

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy import inspect, text
from sqlalchemy.orm import Session

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image,
    PageBreak,
    KeepTogether,
)
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.lineplots import LinePlot
from reportlab.graphics.charts.axes import XValueAxis, YValueAxis

import openpyxl

from app.database import get_db
from app.deps import get_current_user
from app.models.user import User
from app.models.skin_profile import SkinProfile
from app.models.assessment import SkinAssessment
from app.models.routine import SkincareRoutine
from app.models.progress import ProgressLog


router = APIRouter(prefix="/api/reports", tags=["Reports & Export"])


# =========================================================
# THEME
# =========================================================

PRIMARY = colors.HexColor("#6D28D9")
PRIMARY_DARK = colors.HexColor("#4C1D95")
PRIMARY_LIGHT = colors.HexColor("#F5F3FF")
INDIGO = colors.HexColor("#4F46E5")
TEXT = colors.HexColor("#172033")
MUTED = colors.HexColor("#64748B")
BORDER = colors.HexColor("#E2E8F0")
BG = colors.HexColor("#F8FAFC")
WHITE = colors.white
GREEN = colors.HexColor("#16A34A")
GREEN_LIGHT = colors.HexColor("#F0FDF4")
RED = colors.HexColor("#DC2626")
RED_LIGHT = colors.HexColor("#FEF2F2")


# =========================================================
# HELPERS
# =========================================================

def _safe_text(value: Any, default: str = "-") -> str:
    if value is None:
        return default

    if isinstance(value, (list, tuple, set)):
        value = ", ".join(
            str(v) for v in value if v is not None
        )

    text_value = str(value).strip()
    return text_value or default


def _escape(value: Any, default: str = "-") -> str:
    value = _safe_text(value, default)
    return (
        value.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def _fmt_score(value: Any) -> str:
    if value is None:
        return "-"

    try:
        return f"{float(value):.1f}/100"
    except (TypeError, ValueError):
        return "-"


def _fmt_change(value: Any) -> str:
    if value is None:
        return "Not enough data"

    try:
        return f"{float(value):+.1f} points"
    except (TypeError, ValueError):
        return "Not enough data"


def _date(value: Any) -> str:
    if not value:
        return "-"

    if isinstance(value, datetime):
        return value.strftime("%d %b %Y")

    try:
        return datetime.fromisoformat(
            str(value).replace("Z", "+00:00")
        ).strftime("%d %b %Y")
    except Exception:
        return str(value)[:10]


# =========================================================
# DATA
# =========================================================

def _get_photos(db: Session, user_id, days: int):
    """
    Reads the existing progress_photos table without requiring
    a dedicated SQLAlchemy photo model.
    """
    try:
        inspector = inspect(db.bind)
        tables = inspector.get_table_names()

        if "progress_photos" not in tables:
            return []

        columns = {
            col["name"]
            for col in inspector.get_columns("progress_photos")
        }

        required = {
            "user_id",
            "photo_type",
            "photo_url",
            "skin_health_score",
        }

        if not required.issubset(columns):
            return []

        select_cols = [
            "photo_type",
            "photo_url",
            "skin_health_score",
        ]

        if "created_at" in columns:
            select_cols.append("created_at")

        if "id" in columns:
            select_cols.append("id")

        where = "user_id = :user_id"
        params = {"user_id": str(user_id)}

        if "created_at" in columns:
            where += " AND created_at >= :cutoff"
            params["cutoff"] = (
                datetime.utcnow() - timedelta(days=days)
            )

        order = (
            "created_at DESC"
            if "created_at" in columns
            else "photo_type ASC"
        )

        query = text(
            f"""
            SELECT {", ".join(select_cols)}
            FROM progress_photos
            WHERE {where}
            ORDER BY {order}
            """
        )

        rows = db.execute(
            query,
            params,
        ).mappings().all()

        return [dict(row) for row in rows]

    except Exception:
        return []


def _gather_report_data(
    db: Session,
    user: User,
    days: int,
):
    profile = (
        db.query(SkinProfile)
        .filter(SkinProfile.user_id == user.id)
        .first()
    )

    assessment = (
        db.query(SkinAssessment)
        .filter(SkinAssessment.user_id == user.id)
        .order_by(SkinAssessment.created_at.desc())
        .first()
    )

    routine = (
        db.query(SkincareRoutine)
        .filter(SkincareRoutine.user_id == user.id)
        .first()
    )

    cutoff = datetime.utcnow() - timedelta(days=days)

    logs = (
        db.query(ProgressLog)
        .filter(
            ProgressLog.user_id == user.id,
            ProgressLog.log_date >= cutoff,
        )
        .order_by(ProgressLog.log_date.asc())
        .all()
    )

    photos = _get_photos(
        db,
        user.id,
        days,
    )

    return (
        profile,
        assessment,
        routine,
        logs,
        photos,
    )


def _calculate_statistics(logs):
    scores = [
        float(log.skin_health_score)
        for log in logs
        if log.skin_health_score is not None
    ]

    latest = scores[-1] if scores else None
    previous = scores[-2] if len(scores) > 1 else None

    average = (
        sum(scores) / len(scores)
        if scores
        else None
    )

    change = (
        latest - previous
        if latest is not None and previous is not None
        else None
    )

    morning = sum(
        bool(log.routine_followed_morning)
        for log in logs
    )

    evening = sum(
        bool(log.routine_followed_evening)
        for log in logs
    )

    total_slots = len(logs) * 2

    adherence = (
        (morning + evening) / total_slots * 100
        if total_slots
        else 0
    )

    return {
        "records": len(logs),
        "latest": latest,
        "previous": previous,
        "average": average,
        "change": change,
        "morning": morning,
        "evening": evening,
        "adherence": round(adherence),
    }


def _photo_by_type(photos, photo_type):
    matches = [
        photo
        for photo in photos
        if str(
            photo.get("photo_type", "")
        ).lower() == photo_type
    ]

    if not matches:
        return None

    matches.sort(
        key=lambda p: p.get("created_at") or "",
        reverse=True,
    )

    return matches[0]


def _photo_path(photo):
    if not photo:
        return None

    photo_url = photo.get("photo_url")

    if not photo_url:
        return None

    photo_url = str(photo_url)

    if photo_url.startswith("/uploads/"):
        return "/app" + photo_url

    if photo_url.startswith("uploads/"):
        return "/app/" + photo_url

    if os.path.exists(photo_url):
        return photo_url

    return None


# =========================================================
# STYLES
# =========================================================

def _build_styles():
    styles = getSampleStyleSheet()

    styles.add(
        ParagraphStyle(
            name="CoverTitle",
            parent=styles["Title"],
            fontName="Helvetica-Bold",
            fontSize=27,
            leading=31,
            textColor=PRIMARY,
            alignment=TA_CENTER,
            spaceAfter=3 * mm,
        )
    )

    styles.add(
        ParagraphStyle(
            name="CoverSubtitle",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=11,
            leading=15,
            textColor=MUTED,
            alignment=TA_CENTER,
            spaceAfter=5 * mm,
        )
    )

    styles.add(
        ParagraphStyle(
            name="Section",
            parent=styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=12,
            leading=15,
            textColor=TEXT,
            spaceBefore=0,
            spaceAfter=2 * mm,
        )
    )

    styles.add(
        ParagraphStyle(
            name="Body",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=9.3,
            leading=13,
            textColor=TEXT,
            spaceAfter=1.5 * mm,
        )
    )

    styles.add(
        ParagraphStyle(
            name="BodyMuted",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=8.3,
            leading=11.5,
            textColor=MUTED,
        )
    )

    styles.add(
        ParagraphStyle(
            name="Small",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=8.4,
            leading=11.5,
            textColor=TEXT,
        )
    )

    styles.add(
        ParagraphStyle(
            name="SmallBold",
            parent=styles["BodyText"],
            fontName="Helvetica-Bold",
            fontSize=8.4,
            leading=11.5,
            textColor=TEXT,
        )
    )

    styles.add(
        ParagraphStyle(
            name="CardLabel",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=7.8,
            leading=9.5,
            textColor=MUTED,
        )
    )

    styles.add(
        ParagraphStyle(
            name="CardValue",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=15,
            leading=18,
            textColor=TEXT,
        )
    )

    styles.add(
        ParagraphStyle(
            name="RoutineTitle",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=10,
            leading=13,
            textColor=PRIMARY_DARK,
        )
    )

    styles.add(
        ParagraphStyle(
            name="HeroScore",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=30,
            leading=34,
            textColor=PRIMARY,
            alignment=TA_CENTER,
        )
    )

    styles.add(
        ParagraphStyle(
            name="HeroLabel",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=9,
            leading=11,
            textColor=MUTED,
            alignment=TA_CENTER,
        )
    )

    styles.add(
        ParagraphStyle(
            name="HeroChange",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=12,
            leading=15,
            alignment=TA_CENTER,
        )
    )

    return styles


def _section_title(title, styles):
    table = Table(
        [[
            Paragraph(
                _escape(title),
                styles["Section"],
            )
        ]],
        colWidths=[174 * mm],
    )

    table.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, -1),
                    PRIMARY_LIGHT,
                ),
                (
                    "LINEBEFORE",
                    (0, 0),
                    (0, -1),
                    3,
                    PRIMARY,
                ),
                (
                    "BOX",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    BORDER,
                ),
                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    8,
                ),
                (
                    "RIGHTPADDING",
                    (0, 0),
                    (-1, -1),
                    8,
                ),
                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    5,
                ),
                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    4,
                ),
            ]
        )
    )

    return table


def _summary_card(
    label,
    value,
    styles,
    value_color=TEXT,
):
    value_style = ParagraphStyle(
        name=f"Value_{str(label).replace(' ', '_')}",
        parent=styles["CardValue"],
        textColor=value_color,
    )

    return [
        Paragraph(
            _escape(label),
            styles["CardLabel"],
        ),
        Spacer(1, 1.5 * mm),
        Paragraph(
            _escape(value),
            value_style,
        ),
    ]


# =========================================================
# CHART
# =========================================================

def _score_chart(logs):
    usable = [
        (index + 1, float(log.skin_health_score))
        for index, log in enumerate(logs)
        if log.skin_health_score is not None
    ]

    if len(usable) < 2:
        return None

    drawing = Drawing(
        174 * mm,
        55 * mm,
    )

    plot = LinePlot()

    plot.x = 18 * mm
    plot.y = 8 * mm
    plot.width = 146 * mm
    plot.height = 38 * mm

    plot.data = [usable]
    plot.joinedLines = True

    plot.lines[0].strokeColor = PRIMARY
    plot.lines[0].strokeWidth = 2.2

    plot.xValueAxis = XValueAxis()
    plot.xValueAxis.valueMin = 1
    plot.xValueAxis.valueMax = max(
        2,
        len(usable),
    )
    plot.xValueAxis.valueStep = 1
    plot.xValueAxis.labels.fontSize = 6
    plot.xValueAxis.labels.fillColor = MUTED

    plot.yValueAxis = YValueAxis()
    plot.yValueAxis.valueMin = 0
    plot.yValueAxis.valueMax = 100
    plot.yValueAxis.valueStep = 20
    plot.yValueAxis.labels.fontSize = 6
    plot.yValueAxis.labels.fillColor = MUTED

    drawing.add(plot)

    return drawing


# =========================================================
# PHOTO SECTION
# =========================================================

def _make_photo_card(
    title,
    photo,
    styles,
):
    path = _photo_path(photo)

    content = [
        Paragraph(
            f"<b>{_escape(title)}</b>",
            styles["SmallBold"],
        ),
        Spacer(1, 2 * mm),
    ]

    if path and os.path.exists(path):
        try:
            image = Image(
                path,
                width=78 * mm,
                height=65 * mm,
                kind="proportional",
            )
            content.append(image)
        except Exception:
            content.append(
                Paragraph(
                    "Photo could not be embedded.",
                    styles["BodyMuted"],
                )
            )
    else:
        placeholder = Table(
            [[
                Paragraph(
                    "No photo available",
                    styles["BodyMuted"],
                )
            ]],
            colWidths=[78 * mm],
            rowHeights=[65 * mm],
        )

        placeholder.setStyle(
            TableStyle(
                [
                    (
                        "BACKGROUND",
                        (0, 0),
                        (-1, -1),
                        BG,
                    ),
                    (
                        "BOX",
                        (0, 0),
                        (-1, -1),
                        0.5,
                        BORDER,
                    ),
                    (
                        "VALIGN",
                        (0, 0),
                        (-1, -1),
                        "MIDDLE",
                    ),
                    (
                        "ALIGN",
                        (0, 0),
                        (-1, -1),
                        "CENTER",
                    ),
                ]
            )
        )

        content.append(placeholder)

    if photo:
        content.extend(
            [
                Spacer(1, 2 * mm),
                Paragraph(
                    f"<b>AI score:</b> "
                    f"{_escape(_fmt_score(photo.get('skin_health_score')))}"
                    f"<br/><b>Analyzed:</b> "
                    f"{_escape(_date(photo.get('created_at')))}",
                    styles["BodyMuted"],
                ),
            ]
        )

    card = Table(
        [[content]],
        colWidths=[82 * mm],
    )

    card.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, -1),
                    WHITE,
                ),
                (
                    "BOX",
                    (0, 0),
                    (-1, -1),
                    0.7,
                    BORDER,
                ),
                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "TOP",
                ),
                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    6,
                ),
                (
                    "RIGHTPADDING",
                    (0, 0),
                    (-1, -1),
                    6,
                ),
                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    6,
                ),
                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    6,
                ),
            ]
        )
    )

    return card


# =========================================================
# ROUTINE TABLES
# =========================================================

def _routine_rows(items, styles):
    rows = []

    for index, item in enumerate(
        items or [],
        start=1,
    ):
        if not isinstance(item, dict):
            rows.append(
                [
                    str(index),
                    Paragraph(
                        _escape(item),
                        styles["Small"],
                    ),
                    "",
                ]
            )
            continue

        step = item.get("step", index)
        category = item.get("category", "")
        instruction = item.get("instruction", "")
        product = item.get(
            "product_suggestion",
            "",
        )

        detail = _escape(instruction)

        if product:
            detail += (
                "<br/><font color='#64748B'>"
                "Suggested: "
                f"{_escape(product)}"
                "</font>"
            )

        rows.append(
            [
                str(step),
                Paragraph(
                    _escape(category),
                    styles["Small"],
                ),
                Paragraph(
                    detail,
                    styles["Small"],
                ),
            ]
        )

    return rows


def _routine_table(
    title,
    items,
    styles,
):
    rows = [
        [
            Paragraph(
                f"<b>{_escape(title)}</b>",
                styles["SmallBold"],
            ),
            "",
            "",
        ],
        [
            Paragraph(
                "<b>Step</b>",
                styles["SmallBold"],
            ),
            Paragraph(
                "<b>Category</b>",
                styles["SmallBold"],
            ),
            Paragraph(
                "<b>Instruction</b>",
                styles["SmallBold"],
            ),
        ],
    ]

    rows.extend(
        _routine_rows(
            items,
            styles,
        )
    )

    table = Table(
        rows,
        colWidths=[
            14 * mm,
            37 * mm,
            123 * mm,
        ],
        repeatRows=2,
    )

    table.setStyle(
        TableStyle(
            [
                (
                    "SPAN",
                    (0, 0),
                    (-1, 0),
                ),
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    PRIMARY_LIGHT,
                ),
                (
                    "BACKGROUND",
                    (0, 1),
                    (-1, 1),
                    BG,
                ),
                (
                    "BOX",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    BORDER,
                ),
                (
                    "INNERGRID",
                    (0, 1),
                    (-1, -1),
                    0.35,
                    BORDER,
                ),
                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "TOP",
                ),
                (
                    "ALIGN",
                    (0, 1),
                    (0, -1),
                    "CENTER",
                ),
                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    5,
                ),
                (
                    "RIGHTPADDING",
                    (0, 0),
                    (-1, -1),
                    5,
                ),
                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    4,
                ),
                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    4,
                ),
            ]
        )
    )

    return table


def _weekly_table(
    items,
    styles,
):
    rows = [
        [
            Paragraph(
                "<b>Day</b>",
                styles["SmallBold"],
            ),
            Paragraph(
                "<b>Treatment</b>",
                styles["SmallBold"],
            ),
            Paragraph(
                "<b>Purpose</b>",
                styles["SmallBold"],
            ),
        ]
    ]

    for item in items or []:
        if not isinstance(item, dict):
            continue

        rows.append(
            [
                Paragraph(
                    _escape(item.get("day", "-")),
                    styles["Small"],
                ),
                Paragraph(
                    _escape(item.get("treatment", "-")),
                    styles["Small"],
                ),
                Paragraph(
                    _escape(item.get("purpose", "-")),
                    styles["Small"],
                ),
            ]
        )

    table = Table(
        rows,
        colWidths=[
            28 * mm,
            58 * mm,
            88 * mm,
        ],
        repeatRows=1,
    )

    table.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    BG,
                ),
                (
                    "BOX",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    BORDER,
                ),
                (
                    "INNERGRID",
                    (0, 0),
                    (-1, -1),
                    0.35,
                    BORDER,
                ),
                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "TOP",
                ),
                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    5,
                ),
                (
                    "RIGHTPADDING",
                    (0, 0),
                    (-1, -1),
                    5,
                ),
                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    4,
                ),
                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    4,
                ),
            ]
        )
    )

    return table


# =========================================================
# PROGRESS TABLE
# =========================================================

def _progress_table(
    logs,
    styles,
):
    rows = [
        [
            Paragraph("<b>Date</b>", styles["SmallBold"]),
            Paragraph("<b>Morning</b>", styles["SmallBold"]),
            Paragraph("<b>Evening</b>", styles["SmallBold"]),
            Paragraph("<b>Score</b>", styles["SmallBold"]),
            Paragraph("<b>Note</b>", styles["SmallBold"]),
        ]
    ]

    for log in logs:
        rows.append(
            [
                Paragraph(
                    _escape(_date(log.log_date)),
                    styles["Small"],
                ),
                Paragraph(
                    "Completed"
                    if log.routine_followed_morning
                    else "Not recorded",
                    styles["Small"],
                ),
                Paragraph(
                    "Completed"
                    if log.routine_followed_evening
                    else "Not recorded",
                    styles["Small"],
                ),
                Paragraph(
                    _escape(
                        _fmt_score(
                            log.skin_health_score
                        )
                    ),
                    styles["SmallBold"],
                ),
                Paragraph(
                    _escape(
                        log.skin_condition_note
                        or "No note added."
                    ),
                    styles["Small"],
                ),
            ]
        )

    table = Table(
        rows,
        colWidths=[
            28 * mm,
            28 * mm,
            28 * mm,
            25 * mm,
            65 * mm,
        ],
        repeatRows=1,
    )

    table.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    PRIMARY_LIGHT,
                ),
                (
                    "BOX",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    BORDER,
                ),
                (
                    "INNERGRID",
                    (0, 0),
                    (-1, -1),
                    0.35,
                    BORDER,
                ),
                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "TOP",
                ),
                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    4,
                ),
                (
                    "RIGHTPADDING",
                    (0, 0),
                    (-1, -1),
                    4,
                ),
                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    4,
                ),
                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    4,
                ),
            ]
        )
    )

    return table


# =========================================================
# PAGE HEADER / FOOTER
# =========================================================

def _page_header_footer(
    canvas_obj,
    doc,
):
    canvas_obj.saveState()

    width, height = A4

    canvas_obj.setStrokeColor(BORDER)
    canvas_obj.line(
        18 * mm,
        height - 12 * mm,
        width - 18 * mm,
        height - 12 * mm,
    )

    canvas_obj.setFont(
        "Helvetica-Bold",
        7.5,
    )
    canvas_obj.setFillColor(PRIMARY)
    canvas_obj.drawString(
        18 * mm,
        height - 9 * mm,
        "SkinIQ",
    )

    canvas_obj.setFont(
        "Helvetica",
        7,
    )
    canvas_obj.setFillColor(MUTED)
    canvas_obj.drawRightString(
        width - 18 * mm,
        height - 9 * mm,
        "AI-Powered Skin Health Report",
    )

    canvas_obj.setStrokeColor(BORDER)
    canvas_obj.line(
        18 * mm,
        11 * mm,
        width - 18 * mm,
        11 * mm,
    )

    canvas_obj.setFont(
        "Helvetica",
        7,
    )
    canvas_obj.setFillColor(MUTED)

    canvas_obj.drawString(
        18 * mm,
        7 * mm,
        "Generated by SkinIQ",
    )

    canvas_obj.drawRightString(
        width - 18 * mm,
        7 * mm,
        f"Page {doc.page}",
    )

    canvas_obj.restoreState()


# =========================================================
# PROFESSIONAL PDF
# =========================================================

def _create_pdf(
    db: Session,
    current_user: User,
    report_type: str,
    days: int,
):
    (
        profile,
        assessment,
        routine,
        logs,
        photos,
    ) = _gather_report_data(
        db,
        current_user,
        days,
    )

    stats = _calculate_statistics(logs)
    styles = _build_styles()

    before_photo = _photo_by_type(
        photos,
        "before",
    )

    current_photo = _photo_by_type(
        photos,
        "current",
    )

    before_score = (
        float(
            before_photo["skin_health_score"]
        )
        if (
            before_photo
            and before_photo.get(
                "skin_health_score"
            ) is not None
        )
        else None
    )

    current_photo_score = (
        float(
            current_photo["skin_health_score"]
        )
        if (
            current_photo
            and current_photo.get(
                "skin_health_score"
            ) is not None
        )
        else None
    )

    photo_change = (
        current_photo_score - before_score
        if (
            before_score is not None
            and current_photo_score is not None
        )
        else None
    )

    current_score = (
        current_photo_score
        if current_photo_score is not None
        else stats["latest"]
    )

    comparison_change = (
        photo_change
        if photo_change is not None
        else stats["change"]
    )

    # -----------------------------------------------------
    # CONCERNS
    # -----------------------------------------------------

    concerns = []

    if assessment:
        concerns = sorted(
            list(assessment.concerns or []),
            key=lambda concern: (
                concern.priority
                if concern.priority is not None
                else 999
            ),
        )

    # -----------------------------------------------------
    # AI INSIGHT
    # -----------------------------------------------------

    if comparison_change is None:
        insight = (
            "Continue recording your skin health and "
            "routine adherence to build a stronger "
            "progress history."
        )
    elif comparison_change > 5:
        insight = (
            "Your recorded skin-health score shows "
            "meaningful improvement. Continue your "
            "current routine consistently and keep "
            "monitoring your progress."
        )
    elif comparison_change > 0:
        insight = (
            "Your recorded skin-health score is improving. "
            "Maintain routine consistency and continue "
            "tracking changes over time."
        )
    elif comparison_change < 0:
        insight = (
            "The latest recorded score is lower than the "
            "comparison score. Review recent routine "
            "consistency and continue monitoring your "
            "skin response."
        )
    else:
        insight = (
            "Your recorded skin-health score is currently "
            "stable. Continue the routine consistently "
            "and keep tracking progress."
        )

    # -----------------------------------------------------
    # DOCUMENT
    # -----------------------------------------------------

    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=18 * mm,
        leftMargin=18 * mm,
        topMargin=17 * mm,
        bottomMargin=17 * mm,
        title=(
            f"SkinIQ {report_type} "
            "Skin Health Report"
        ),
        author="SkinIQ",
    )

    story = []

    # =====================================================
    # PAGE 1 - EXECUTIVE SUMMARY
    # =====================================================

    story.append(Spacer(1, 8 * mm))

    story.append(
        Paragraph(
            "SkinIQ",
            styles["CoverTitle"],
        )
    )

    story.append(
        Paragraph(
            "AI-Powered Skin Health Report",
            styles["CoverSubtitle"],
        )
    )

    metadata = Table(
        [[
            Paragraph(
                "<b>Report Type</b><br/>"
                + _escape(report_type),
                styles["Small"],
            ),
            Paragraph(
                "<b>Generated</b><br/>"
                + datetime.now().strftime(
                    "%d %B %Y"
                ),
                styles["Small"],
            ),
            Paragraph(
                "<b>Period</b><br/>"
                + f"Last {days} days",
                styles["Small"],
            ),
        ]],
        colWidths=[
            58 * mm,
            58 * mm,
            58 * mm,
        ],
    )

    metadata.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), BG),
                ("BOX", (0, 0), (-1, -1), 0.6, BORDER),
                ("INNERGRID", (0, 0), (-1, -1), 0.35, BORDER),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 7),
                ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )

    story.append(metadata)
    story.append(Spacer(1, 5 * mm))

    # HERO SCORE

    score_display = _fmt_score(current_score)

    change_color = (
        GREEN
        if comparison_change is not None
        and comparison_change >= 0
        else RED
    )

    change_display = (
        _fmt_change(comparison_change)
        if comparison_change is not None
        else "No comparison yet"
    )

    hero = Table(
        [[
            [
                Paragraph(
                    "CURRENT SKIN HEALTH SCORE",
                    styles["HeroLabel"],
                ),
                Spacer(1, 2 * mm),
                Paragraph(
                    _escape(score_display),
                    styles["HeroScore"],
                ),
                Spacer(1, 2 * mm),
                Paragraph(
                    _escape(change_display),
                    ParagraphStyle(
                        "HeroChangeColor",
                        parent=styles["HeroChange"],
                        textColor=change_color,
                    ),
                ),
            ]
        ]],
        colWidths=[174 * mm],
    )

    hero.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, -1),
                    PRIMARY_LIGHT,
                ),
                (
                    "BOX",
                    (0, 0),
                    (-1, -1),
                    0.8,
                    colors.HexColor("#DDD6FE"),
                ),
                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "MIDDLE",
                ),
                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    10,
                ),
                (
                    "RIGHTPADDING",
                    (0, 0),
                    (-1, -1),
                    10,
                ),
                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    8,
                ),
                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    8,
                ),
            ]
        )
    )

    story.append(hero)
    story.append(Spacer(1, 3 * mm))

    # USER PROFILE

    story.append(
        _section_title(
            "1. User & Skin Profile",
            styles,
        )
    )

    profile_rows = [
        [
            "Name",
            _safe_text(
                current_user.full_name
            ),
        ],
        [
            "Email",
            _safe_text(
                current_user.email
            ),
        ],
        [
            "Skin Type",
            (
                _safe_text(
                    getattr(
                        profile,
                        "skin_type",
                        None,
                    )
                )
                if profile
                else "-"
            ),
        ],
        [
            "Age Group",
            (
                _safe_text(
                    getattr(
                        profile,
                        "age_group",
                        None,
                    )
                )
                if profile
                else "-"
            ),
        ],
        [
            "Skin Concerns",
            (
                _safe_text(
                    getattr(
                        profile,
                        "skin_concerns",
                        None,
                    )
                )
                if profile
                else "-"
            ),
        ],
    ]

    profile_table = Table(
        [
            [
                Paragraph(
                    f"<b>{_escape(label)}</b>",
                    styles["Small"],
                ),
                Paragraph(
                    _escape(value),
                    styles["Small"],
                ),
            ]
            for label, value in profile_rows
        ],
        colWidths=[
            40 * mm,
            134 * mm,
        ],
    )

    profile_table.setStyle(
        TableStyle(
            [
                ("BOX", (0, 0), (-1, -1), 0.5, BORDER),
                ("INNERGRID", (0, 0), (-1, -1), 0.35, BORDER),
                ("BACKGROUND", (0, 0), (0, -1), BG),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )

    story.append(profile_table)
    story.append(Spacer(1, 3 * mm))

    # SUMMARY CARDS

    story.append(
        _section_title(
            "2. Progress Summary",
            styles,
        )
    )

    summary = Table(
        [
            [
                _summary_card(
                    "Average Score",
                    _fmt_score(stats["average"]),
                    styles,
                    PRIMARY,
                ),
                _summary_card(
                    "Progress Records",
                    str(stats["records"]),
                    styles,
                ),
                _summary_card(
                    "Routine Adherence",
                    f"{stats['adherence']}%",
                    styles,
                    GREEN,
                ),
            ],
            [
                _summary_card(
                    "Morning Completed",
                    str(stats["morning"]),
                    styles,
                ),
                _summary_card(
                    "Evening Completed",
                    str(stats["evening"]),
                    styles,
                ),
                _summary_card(
                    "Photo Comparison",
                    (
                        f"{before_score:.1f} -> "
                        f"{current_photo_score:.1f}"
                        if (
                            before_score is not None
                            and current_photo_score is not None
                        )
                        else "Not available"
                    ),
                    styles,
                ),
            ],
        ],
        colWidths=[
            58 * mm,
            58 * mm,
            58 * mm,
        ],
        rowHeights=[
            19 * mm,
            19 * mm,
        ],
    )

    summary.setStyle(
        TableStyle(
            [
                ("BOX", (0, 0), (-1, -1), 0.5, BORDER),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, BORDER),
                ("BACKGROUND", (0, 0), (-1, -1), WHITE),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 7),
                ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )

    story.append(summary)
    story.append(Spacer(1, 4 * mm))

    # AI ASSESSMENT

    story.append(
        _section_title(
            "3. AI Skin Assessment",
            styles,
        )
    )

    if assessment:
        assessment_items = [
            [
                Paragraph(
                    "<b>Assessment Score</b>",
                    styles["Small"],
                ),
                Paragraph(
                    _escape(
                        _fmt_score(
                            assessment.condition_score
                        )
                    ),
                    styles["SmallBold"],
                ),
            ]
        ]

        for concern in concerns:
            assessment_items.append(
                [
                    Paragraph(
                        _escape(
                            concern.concern_name
                        ),
                        styles["Small"],
                    ),
                    Paragraph(
                        "Severity: "
                        + _escape(
                            concern.severity
                        )
                        + " | Priority: "
                        + _escape(
                            concern.priority
                        ),
                        styles["Small"],
                    ),
                ]
            )

        assessment_table = Table(
            assessment_items,
            colWidths=[
                55 * mm,
                119 * mm,
            ],
        )

        assessment_table.setStyle(
            TableStyle(
                [
                    ("BOX", (0, 0), (-1, -1), 0.5, BORDER),
                    ("INNERGRID", (0, 0), (-1, -1), 0.35, BORDER),
                    ("BACKGROUND", (0, 0), (0, -1), BG),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 5),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )

        story.append(assessment_table)
    else:
        story.append(
            Paragraph(
                "No skin assessment has been recorded.",
                styles["Body"],
            )
        )

    # =====================================================
    # BEFORE / CURRENT + AI INSIGHT
    # =====================================================
    # Section 4 is intentionally kept together below so its
    # heading never appears alone at the bottom of a page.

    story.append(Spacer(1, 4 * mm))

    photo_table = Table(
        [[
            _make_photo_card(
                "BEFORE",
                before_photo,
                styles,
            ),
            _make_photo_card(
                "CURRENT",
                current_photo,
                styles,
            ),
        ]],
        colWidths=[
            85 * mm,
            85 * mm,
        ],
    )

    photo_table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )

    # The photo table is rendered only inside Section 4 below.
    # Do not append it here separately, otherwise the Before/Current
    # photos appear twice in the PDF.

    if photo_change is not None:
        comparison_color = (
            GREEN
            if photo_change >= 0
            else RED
        )

        comparison_box = Table(
            [[
                Paragraph(
                    "PHOTO PROGRESS",
                    styles["HeroLabel"],
                ),
                Paragraph(
                    f"{before_score:.1f}/100",
                    styles["CardValue"],
                ),
                Paragraph(
                    "→",
                    styles["HeroChange"],
                ),
                Paragraph(
                    f"{current_photo_score:.1f}/100",
                    styles["CardValue"],
                ),
                Paragraph(
                    f"{photo_change:+.1f} points",
                    ParagraphStyle(
                        "ComparisonChange",
                        parent=styles["HeroChange"],
                        fontSize=16,
                        leading=19,
                        textColor=comparison_color,
                    ),
                ),
            ]],
            colWidths=[
                34 * mm,
                34 * mm,
                18 * mm,
                34 * mm,
                54 * mm,
            ],
        )

        comparison_box.setStyle(
            TableStyle(
                [
                    (
                        "BACKGROUND",
                        (0, 0),
                        (-1, -1),
                        GREEN_LIGHT
                        if photo_change >= 0
                        else colors.HexColor("#FEF2F2"),
                    ),
                    (
                        "BOX",
                        (0, 0),
                        (-1, -1),
                        0.7,
                        BORDER,
                    ),
                    (
                        "VALIGN",
                        (0, 0),
                        (-1, -1),
                        "MIDDLE",
                    ),
                    (
                        "ALIGN",
                        (0, 0),
                        (-1, -1),
                        "CENTER",
                    ),
                    (
                        "LEFTPADDING",
                        (0, 0),
                        (-1, -1),
                        5,
                    ),
                    (
                        "RIGHTPADDING",
                        (0, 0),
                        (-1, -1),
                        5,
                    ),
                    (
                        "TOPPADDING",
                        (0, 0),
                        (-1, -1),
                        7,
                    ),
                    (
                        "BOTTOMPADDING",
                        (0, 0),
                        (-1, -1),
                        7,
                    ),
                ]
            )
        )
    else:
        comparison_box = Table(
            [[
                Paragraph(
                    "<b>PHOTO PROGRESS</b><br/>"
                    "Upload both Before and Current photos "
                    "to calculate your photo-based improvement.",
                    styles["Body"],
                )
            ]],
            colWidths=[174 * mm],
        )

        comparison_box.setStyle(
            TableStyle(
                [
                    (
                        "BACKGROUND",
                        (0, 0),
                        (-1, -1),
                        BG,
                    ),
                    (
                        "BOX",
                        (0, 0),
                        (-1, -1),
                        0.5,
                        BORDER,
                    ),
                    (
                        "LEFTPADDING",
                        (0, 0),
                        (-1, -1),
                        8,
                    ),
                    (
                        "RIGHTPADDING",
                        (0, 0),
                        (-1, -1),
                        8,
                    ),
                    (
                        "TOPPADDING",
                        (0, 0),
                        (-1, -1),
                        8,
                    ),
                    (
                        "BOTTOMPADDING",
                        (0, 0),
                        (-1, -1),
                        8,
                    ),
                ]
            )
        )

    comparison_box.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, -1),
                    GREEN_LIGHT
                    if photo_change is not None
                    and photo_change >= 0
                    else BG,
                ),
                (
                    "BOX",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    BORDER,
                ),
                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    8,
                ),
                (
                    "RIGHTPADDING",
                    (0, 0),
                    (-1, -1),
                    8,
                ),
                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    7,
                ),
                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    7,
                ),
            ]
        )
    )

    # Keep the Section 4 heading attached to its photo comparison.
    # This prevents the heading from being stranded at the bottom
    # of the previous page when the photo block moves to the next page.
    story.append(
        KeepTogether(
            [
                Spacer(1, 4 * mm),
                _section_title(
                    "4. Before & Current Comparison",
                    styles,
                ),
                photo_table,
                Spacer(1, 3 * mm),
                comparison_box,
            ]
        )
    )

    story.append(Spacer(1, 7 * mm))

    story.append(
        _section_title(
            "5. AI Progress Analysis",
            styles,
        )
    )

    insight_box = Table(
        [[
            Paragraph(
                _escape(insight),
                styles["Body"],
            )
        ]],
        colWidths=[174 * mm],
    )

    insight_box.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, -1),
                    PRIMARY_LIGHT,
                ),
                (
                    "BOX",
                    (0, 0),
                    (-1, -1),
                    0.6,
                    colors.HexColor("#DDD6FE"),
                ),
                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    9,
                ),
                (
                    "RIGHTPADDING",
                    (0, 0),
                    (-1, -1),
                    9,
                ),
                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    8,
                ),
                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    8,
                ),
            ]
        )
    )

    story.append(insight_box)

    # =====================================================
    # PERSONALIZED ROUTINE
    # =====================================================

    story.append(Spacer(1, 5 * mm))

    story.append(
        _section_title(
            "6. Personalized Skincare Routine",
            styles,
        )
    )

    if routine:
        season = getattr(
            routine,
            "season",
            "all",
        )

        notes = getattr(
            routine,
            "notes",
            "",
        )

        story.append(
            Paragraph(
                f"<b>Season:</b> "
                f"{_escape(season)}",
                styles["Body"],
            )
        )

        if notes:
            story.append(
                Paragraph(
                    f"<b>Routine guidance:</b> "
                    f"{_escape(notes)}",
                    styles["Body"],
                )
            )

        morning = (
            getattr(
                routine,
                "morning_routine",
                None,
            )
            or []
        )

        evening = (
            getattr(
                routine,
                "evening_routine",
                None,
            )
            or []
        )

        weekly = (
            getattr(
                routine,
                "weekly_treatments",
                None,
            )
            or []
        )

        if morning:
            story.append(Spacer(1, 3 * mm))
            story.append(
                _routine_table(
                    "MORNING ROUTINE  |  DAILY",
                    morning,
                    styles,
                )
            )

        if evening:
            story.append(Spacer(1, 6 * mm))
            story.append(
                _routine_table(
                    "EVENING ROUTINE  |  DAILY",
                    evening,
                    styles,
                )
            )

        if weekly:
            story.append(Spacer(1, 7 * mm))
            story.append(
                Paragraph(
                    "WEEKLY TREATMENTS",
                    styles["RoutineTitle"],
                )
            )
            story.append(Spacer(1, 2 * mm))
            story.append(
                _weekly_table(
                    weekly,
                    styles,
                )
            )

    else:
        story.append(
            Paragraph(
                "No personalized skincare routine has been recorded.",
                styles["Body"],
            )
        )

    # =====================================================
    # PROGRESS HISTORY
    # =====================================================

    story.append(Spacer(1, 5 * mm))

    story.append(
        _section_title(
            "7. Progress History",
            styles,
        )
    )

    story.append(
        Paragraph(
            f"Daily records included in this "
            f"{report_type.lower()} report.",
            styles["BodyMuted"],
        )
    )

    if logs:
        story.append(Spacer(1, 3 * mm))
        story.append(
            _progress_table(
                logs,
                styles,
            )
        )

        chart = _score_chart(logs)

        if chart:
            story.append(Spacer(1, 6 * mm))
            story.append(
                Paragraph(
                    "Skin Health Score Trend",
                    styles["RoutineTitle"],
                )
            )
            story.append(
                Paragraph(
                    "Recorded scores across the selected report period.",
                    styles["BodyMuted"],
                )
            )
            story.append(chart)
    else:
        story.append(
            Spacer(1, 4 * mm)
        )
        story.append(
            Paragraph(
                "No progress records were found for this period.",
                styles["Body"],
            )
        )

    story.append(Spacer(1, 7 * mm))

    story.append(
        _section_title(
            "8. Professional Progress Summary",
            styles,
        )
    )

    final_summary = (
        f"During the last {days} days, SkinIQ recorded "
        f"{stats['records']} progress record(s), with "
        f"{stats['adherence']}% routine adherence. "
        f"The latest available score is "
        f"{_fmt_score(current_score)}. "
        f"{insight}"
    )

    summary_box = Table(
        [[
            Paragraph(
                _escape(final_summary),
                styles["Body"],
            )
        ]],
        colWidths=[174 * mm],
    )

    summary_box.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, -1),
                    BG,
                ),
                (
                    "BOX",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    BORDER,
                ),
                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    9,
                ),
                (
                    "RIGHTPADDING",
                    (0, 0),
                    (-1, -1),
                    9,
                ),
                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    8,
                ),
                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    8,
                ),
            ]
        )
    )

    story.append(summary_box)

    story.append(Spacer(1, 7 * mm))

    disclaimer = Table(
        [[
            Paragraph(
                "<b>Important:</b> This report is generated "
                "from recorded skincare information and "
                "AI-assisted analysis. It is intended for "
                "personal progress tracking and informational "
                "purposes only and is not a medical diagnosis.",
                styles["BodyMuted"],
            )
        ]],
        colWidths=[174 * mm],
    )

    disclaimer.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, -1),
                    BG,
                ),
                (
                    "BOX",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    BORDER,
                ),
                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    8,
                ),
                (
                    "RIGHTPADDING",
                    (0, 0),
                    (-1, -1),
                    8,
                ),
                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    7,
                ),
                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    7,
                ),
            ]
        )
    )

    story.append(disclaimer)

    # BUILD PDF

    doc.build(
        story,
        onFirstPage=_page_header_footer,
        onLaterPages=_page_header_footer,
    )

    buffer.seek(0)
    return buffer


# =========================================================
# WEEKLY PDF
# =========================================================

@router.get("/weekly/pdf")
def export_weekly_pdf(
    db: Session = Depends(get_db),
    current_user: User = Depends(
        get_current_user
    ),
):
    buffer = _create_pdf(
        db=db,
        current_user=current_user,
        report_type="Weekly",
        days=7,
    )

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": (
                'attachment; '
                'filename="skiniq_weekly_skin_health_report.pdf"'
            )
        },
    )


# =========================================================
# MONTHLY PDF
# =========================================================

@router.get("/monthly/pdf")
def export_monthly_pdf(
    db: Session = Depends(get_db),
    current_user: User = Depends(
        get_current_user
    ),
):
    buffer = _create_pdf(
        db=db,
        current_user=current_user,
        report_type="Monthly",
        days=30,
    )

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": (
                'attachment; '
                'filename="skiniq_monthly_skin_health_report.pdf"'
            )
        },
    )


# =========================================================
# EXCEL EXPORT
# =========================================================

@router.get("/excel")
def export_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(
        get_current_user
    ),
):
    (
        profile,
        assessment,
        routine,
        logs,
        photos,
    ) = _gather_report_data(
        db,
        current_user,
        days=30,
    )

    wb = openpyxl.Workbook()

    # -----------------------------------------------------
    # SUMMARY
    # -----------------------------------------------------

    ws = wb.active
    ws.title = "Summary"

    stats = _calculate_statistics(logs)

    before_photo = _photo_by_type(
        photos,
        "before",
    )

    current_photo = _photo_by_type(
        photos,
        "current",
    )

    ws.append([
        "SkinIQ Professional Skin Health Report"
    ])

    ws.append([])

    ws.append([
        "Name",
        current_user.full_name,
    ])

    ws.append([
        "Email",
        current_user.email,
    ])

    ws.append([
        "Skin Type",
        (
            getattr(
                profile,
                "skin_type",
                "-"
            )
            if profile
            else "-"
        ),
    ])

    ws.append([
        "Age Group",
        (
            getattr(
                profile,
                "age_group",
                "-"
            )
            if profile
            else "-"
        ),
    ])

    ws.append([
        "Current Score",
        stats["latest"],
    ])

    ws.append([
        "Average Score",
        stats["average"],
    ])

    ws.append([
        "Score Change",
        stats["change"],
    ])

    ws.append([
        "Routine Adherence (%)",
        stats["adherence"],
    ])

    ws.append([
        "Before Photo Score",
        (
            before_photo.get(
                "skin_health_score"
            )
            if before_photo
            else None
        ),
    ])

    ws.append([
        "Current Photo Score",
        (
            current_photo.get(
                "skin_health_score"
            )
            if current_photo
            else None
        ),
    ])

    # -----------------------------------------------------
    # PROGRESS LOG
    # -----------------------------------------------------

    ws2 = wb.create_sheet(
        "Progress Log"
    )

    ws2.append([
        "Date",
        "Morning Followed",
        "Evening Followed",
        "Skin Health Score",
        "Note",
    ])

    for log in logs:
        ws2.append([
            log.log_date.strftime(
                "%Y-%m-%d"
            ),
            bool(
                log.routine_followed_morning
            ),
            bool(
                log.routine_followed_evening
            ),
            (
                float(
                    log.skin_health_score
                )
                if log.skin_health_score
                is not None
                else None
            ),
            log.skin_condition_note or "",
        ])

    # -----------------------------------------------------
    # ASSESSMENT
    # -----------------------------------------------------

    ws3 = wb.create_sheet(
        "Assessment"
    )

    ws3.append([
        "Assessment Score",
        "Concern",
        "Severity",
        "Priority",
    ])

    if assessment:
        concerns = sorted(
            list(
                assessment.concerns or []
            ),
            key=lambda concern: (
                concern.priority
                if concern.priority is not None
                else 999
            ),
        )

        if concerns:
            for concern in concerns:
                ws3.append([
                    (
                        float(
                            assessment.condition_score
                        )
                        if assessment.condition_score
                        is not None
                        else None
                    ),
                    concern.concern_name,
                    concern.severity,
                    concern.priority,
                ])
        else:
            ws3.append([
                (
                    float(
                        assessment.condition_score
                    )
                    if assessment.condition_score
                    is not None
                    else None
                ),
                "No concerns recorded",
                "",
                "",
            ])

    # -----------------------------------------------------
    # ROUTINE
    # -----------------------------------------------------

    ws4 = wb.create_sheet(
        "Routine"
    )

    ws4.append([
        "Time",
        "Step",
        "Category / Day",
        "Instruction / Treatment",
        "Suggested Product / Purpose",
    ])

    if routine:
        for item in (
            getattr(
                routine,
                "morning_routine",
                None,
            )
            or []
        ):
            if isinstance(item, dict):
                ws4.append([
                    "Morning",
                    item.get("step", ""),
                    item.get("category", ""),
                    item.get("instruction", ""),
                    item.get(
                        "product_suggestion",
                        "",
                    ),
                ])

        for item in (
            getattr(
                routine,
                "evening_routine",
                None,
            )
            or []
        ):
            if isinstance(item, dict):
                ws4.append([
                    "Evening",
                    item.get("step", ""),
                    item.get("category", ""),
                    item.get("instruction", ""),
                    item.get(
                        "product_suggestion",
                        "",
                    ),
                ])

        for item in (
            getattr(
                routine,
                "weekly_treatments",
                None,
            )
            or []
        ):
            if isinstance(item, dict):
                ws4.append([
                    "Weekly",
                    "",
                    item.get("day", ""),
                    item.get("treatment", ""),
                    item.get("purpose", ""),
                ])

    # Formatting

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"

        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(
                    vertical="top",
                    wrap_text=True,
                )

        for column_cells in sheet.columns:
            max_length = 0

            column_letter = (
                column_cells[0].column_letter
            )

            for cell in column_cells:
                value = (
                    ""
                    if cell.value is None
                    else str(cell.value)
                )

                max_length = max(
                    max_length,
                    len(value),
                )

            sheet.column_dimensions[
                column_letter
            ].width = min(
                max(
                    max_length + 2,
                    12,
                ),
                45,
            )

    buffer = io.BytesIO()

    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": (
                'attachment; '
                'filename="skiniq_progress_report.xlsx"'
            )
        },
    )
